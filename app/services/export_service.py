"""
Export service — generates JSON, CSV, and Excel files from scraped results.
Files are stored in the exports/ directory and tracked in the database.
"""
from __future__ import annotations

import csv
import json
import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.models import db, ScrapeJob, ScrapeResult, ExportRecord

logger = logging.getLogger(__name__)

EXPORT_DIR = Path("exports")


def _ensure_export_dir() -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    return EXPORT_DIR


def _get_results(job_id: int) -> list[ScrapeResult]:
    return (
        ScrapeResult.query.filter_by(job_id=job_id)
        .order_by(ScrapeResult.page_num, ScrapeResult.item_index)
        .all()
    )


def _record_export(job_id: int, fmt: str, filename: str, filepath: str, row_count: int) -> ExportRecord:
    size = os.path.getsize(filepath) if os.path.exists(filepath) else 0
    record = ExportRecord(
        job_id=job_id,
        format=fmt,
        filename=filename,
        filepath=filepath,
        file_size_bytes=size,
        row_count=row_count,
    )
    db.session.add(record)
    db.session.commit()
    return record


def export_json(job_id: int) -> Optional[str]:
    """Export results as a JSON file. Returns the filepath."""
    job = ScrapeJob.query.get(job_id)
    if not job:
        return None

    results = _get_results(job_id)
    export_dir = _ensure_export_dir()
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"job_{job_id}_{ts}.json"
    filepath = str(export_dir / filename)

    data = {
        "job": job.to_dict(),
        "total_items": len(results),
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "results": [r.to_dict() for r in results],
    }

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    _record_export(job_id, "json", filename, filepath, len(results))
    logger.info("JSON export for job %s: %s", job_id, filepath)
    return filepath


def export_csv(job_id: int) -> Optional[str]:
    """Export results as a CSV file."""
    job = ScrapeJob.query.get(job_id)
    if not job:
        return None

    results = _get_results(job_id)
    export_dir = _ensure_export_dir()
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"job_{job_id}_{ts}.csv"
    filepath = str(export_dir / filename)

    fieldnames = ["id", "job_id", "page_num", "item_index", "page_url", "content_type", "content", "created_at"]

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:  # BOM for Excel compat
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for r in results:
            row = r.to_dict()
            row.pop("metadata", None)
            writer.writerow(row)

    _record_export(job_id, "csv", filename, filepath, len(results))
    logger.info("CSV export for job %s: %s", job_id, filepath)
    return filepath


def export_excel(job_id: int) -> Optional[str]:
    """Export results as a styled Excel (.xlsx) file."""
    job = ScrapeJob.query.get(job_id)
    if not job:
        return None

    results = _get_results(job_id)
    export_dir = _ensure_export_dir()
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"job_{job_id}_{ts}.xlsx"
    filepath = str(export_dir / filename)

    wb = openpyxl.Workbook()

    # --- Results sheet ---
    ws = wb.active
    ws.title = "Results"

    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["#", "Page", "Index", "URL", "Type", "Content", "Scraped At"]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r.id)
        ws.cell(row=row_idx, column=2, value=r.page_num)
        ws.cell(row=row_idx, column=3, value=r.item_index)
        ws.cell(row=row_idx, column=4, value=r.page_url)
        ws.cell(row=row_idx, column=5, value=r.content_type)
        ws.cell(row=row_idx, column=6, value=(r.content or "")[:32767])  # Excel cell limit
        ws.cell(row=row_idx, column=7, value=r.created_at.isoformat() if r.created_at else "")

    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["F"].width = 60

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary")
    job_dict = job.to_dict()
    ws2.cell(1, 1, "Field").font = Font(bold=True)
    ws2.cell(1, 2, "Value").font = Font(bold=True)
    for i, (k, v) in enumerate(job_dict.items(), 2):
        ws2.cell(i, 1, k)
        ws2.cell(i, 2, str(v))
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 50

    wb.save(filepath)
    _record_export(job_id, "excel", filename, filepath, len(results))
    logger.info("Excel export for job %s: %s", job_id, filepath)
    return filepath


def get_job_exports(job_id: int) -> list[ExportRecord]:
    return ExportRecord.query.filter_by(job_id=job_id).order_by(ExportRecord.created_at.desc()).all()
